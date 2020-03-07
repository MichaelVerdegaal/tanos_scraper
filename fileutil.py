from openpyxl import Workbook, load_workbook
from os.path import exists

filename = "vocab.xlsx"


def check_workbook():
    """Creates an excel file if it doesn't exist already"""
    if exists(filename):
        pass
    else:
        wb = Workbook()
        wb.save(filename)
        wb.close()


def return_workbook():
    """
    Loads an openpyxl workbook for the relevant excel file
    :return: workbook
    """
    check_workbook()
    workbook = load_workbook(filename)
    return workbook
